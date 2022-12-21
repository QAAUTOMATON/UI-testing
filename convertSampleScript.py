import time
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

inputSheetName = 'convertSampleScript'
inputFileName = 'conversionStudio.xlsx'
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()


# =======================
def wait(n):
    return time.sleep(n / 4)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element):
    if element.is_displayed():
        element.clear()
        element.send_keys("@#$%!@#$%^&*")
        element.send_keys("QWERTYUqwerty")
        element.send_keys("1234567890")
        wait(1)
        return 1



def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)

def input_limit(element):
    count = 0
    for i in range(0, 62):
        element.send_keys("abcdefghijklmnop")
        count += 16
    if count < 1000:
        return 0
    else:
        return 1

# ===================================================
current_row = 6
current_column = 18
# -------------------------Logging in-------------------------------
login()

current_row = 6
try:
    ele = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[6]/div/button')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        ele2 = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[6]/div/div/a[2]')
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

current_row = 8
try:
    ele = driver.find_element(By.ID, 'actionline_select')
    if ele.is_displayed():
        ele.click()
        dropdown_test(ele)
        Select(ele).select_by_visible_text("Hindi")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 9
try:
    ele = driver.find_element(By.ID, 'actionline_select')
    if ele.is_displayed():
        ele.click()
        dropdown_test(ele)
        Select(ele).select_by_visible_text("English")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 10
try:
    ele = driver.find_element(By.ID, 'actionline_select')
    if ele.is_displayed():
        ele.click()
        dropdown_test(ele)
        Select(ele).select_by_visible_text("Bengali")
        wait(10)
        Select(ele).select_by_index(0)
        wait(20)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 11
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


current_row = 12
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 13
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[3]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 14
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[4]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


current_row = 15
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[5]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 16
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[6]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 17
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[7]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 18
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[8]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 19
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[9]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 20
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[10]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 21
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[11]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 22
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[12]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 23
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[13]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 24
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[14]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 25
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[15]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 26
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[16]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 27
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[17]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10)
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 28
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[18]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 29
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[19]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 30
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[20]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 31
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[21]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 32
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[22]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 33
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[23]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 34
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[24]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 35
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[25]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 36
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[26]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 37
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[27]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 38
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[28]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 39
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[29]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 40
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[30]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 41
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[31]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 42
try:
    ele = driver.find_element(By.XPATH, '//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[32]/div[1]/a[2]')
    if ele.is_displayed():
        ele.click()
        wait(10);
        driver.back()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

