import time
from datetime import date

import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'BecomeLPP'
inputFileName = 'OpportunitiesWithMNF.xlsx'

pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

day = date.today()
# ==========================================================================================================
path = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(path)

url = 'https://mynextfilm.com/members-home'
driver.get(url)

driver.maximize_window()

# =======================
def wait(n):
    return time.sleep(n / 10)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


# -------------------------------------------------------------------------
def input_box_by_name(name, keys):
    ele = driver.find_element(By.NAME, name)
    ele.clear()
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        ele.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1
# ===================================================
current_row = 5
current_column = 18
# -------------------------Logging in-------------------------------
login()
# -------------------------Navigation to Opp With MNF option-------------------------------
current_row = 5
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              'body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(2) div.dropdown > button.dropbtn1.opport-w-m')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Navigation to Opp With MNF: LPP option-------------------------------
current_row = 6
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'Become Language Pair Partner')]")
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box FirstName-------------------------------
current_row = 7
try:
    if input_box_by_name('firstName', 'Firstname') == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box LastName-------------------------------
current_row = 8
try:
    if input_box_by_name('lastName', 'lastName') == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box Mail-------------------------------
current_row = 9
try:
    ele = driver.find_element(By.NAME, 'emailId')
    if ele.is_displayed():
        ele.clear()
        ele.send_keys("testMail@mail.com")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

current_row = 10
try:
    ele = driver.find_element(By.NAME, 'emailId')
    if ele.is_displayed():
        ele.clear()
        ele.send_keys("testMail@mail.com")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)
# -------------------------Dropdown Country code Selection-------------------------------
current_row = 11
try:
    ele = driver.find_element(By.ID, 'code')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box phone no-------------------------------
current_row = 12
try:
    ele = driver.find_element(By.NAME, 'number')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        if len(ele.get_attribute('value')) == 0:
            ele.send_keys("9543219876")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown Country Selection-------------------------------
current_row = 13
try:
    ele = driver.find_element(By.ID, 'country')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown region Selection-------------------------------
current_row = 15
try:
    ele = driver.find_element(By.ID, 'region')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown Country Selection-------------------------------
current_row = 17
try:
    ele = driver.find_element(By.ID, 'city')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown mother tongue Selection-------------------------------
current_row = 18
try:
    ele = driver.find_element(By.ID, 'selectNumber')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown fatherTongue Selection-------------------------------
current_row = 20
try:
    ele = driver.find_element(By.ID, 'fatherTongue')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown motherTongue Selection-------------------------------
current_row = 22
try:
    ele = driver.find_element(By.ID, 'motherTongue')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------checkboxSelection-------------------------------
try:
    ele = driver.find_element(By.XPATH,"//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/form[1]/div[3]/p[1]/input[1]")
    ele.click()
except:
    print("checkbox opened")

body = driver.find_element(By.CSS_SELECTOR, 'body')
body.send_keys(Keys.PAGE_DOWN)
wait(2)
# -------------------------Dropdown firstLang Selection-------------------------------
current_row = 24
try:
    ele = driver.find_element(By.ID, 'firstLanguage')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown firstLang Selection-------------------------------
current_row = 26
try:
    ele = driver.find_element(By.ID, 'motherTounge')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown firstLang Selection-------------------------------
current_row = 27
try:
    ele = driver.find_element(By.ID, 'nativLanguage')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown firstLang Selection-------------------------------
current_row = 28
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction')
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele = driver.find_element(By.NAME, 'mediumOfInstructionL1')
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown secondLang Selection-------------------------------
current_row = 29
try:
    ele = driver.find_element(By.NAME, 'secondLanguage')
    if ele.is_displayed():
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown secondLang Selection-------------------------------
current_row = 31
try:
    ele = driver.find_element(By.ID, 'motherTounge1')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown secondLang Selection-------------------------------
current_row = 32
try:
    ele = driver.find_element(By.ID, 'nativLanguage1')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown secondLang Selection-------------------------------
current_row = 33
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction1')
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele = driver.find_element(By.NAME, 'mediumOfInstructionL2')
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 2
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 2
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------Dropdown secondLang Selection-------------------------------
current_row = 35
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction1')
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele.click()
        wait(1)
        ele = driver.find_element(By.NAME, 'mediumOfInstructionL2')
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 2
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 2
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------checkboxSelection-------------------------------
try:
    ele = driver.find_element(By.XPATH,"//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/form[1]/div[4]/div[8]/p[1]/input[1]")
    ele.click()
except:
    print("checkbox opened")

body = driver.find_element(By.CSS_SELECTOR, 'body')
body.send_keys(Keys.PAGE_DOWN)
wait(2)
# -------------------------Dropdown qualification Selection-------------------------------
current_row = 36
try:
    ele = driver.find_element(By.NAME, 'firstLangDegree')
    if ele.is_displayed():
        wait(1)
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 2
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 2
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box FirstName-------------------------------
current_row = 38
try:
    ele = driver.find_element(By.NAME, 'firstLangAwardedBy')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        if len(ele.get_attribute('value')) == 0:
            ele.send_keys("sample uni")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------File upload qualification Selection-------------------------------
current_row = 39
try:
    ele = driver.find_element(By.ID, 'myfile1')
    if ele.is_displayed():
        wait(1)
        ele.send_keys(pdf_path)
        for i in range(0, 2):
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)
# -------------------------Dropdown qualification Selection-------------------------------
current_row = 41
try:
    ele = driver.find_element(By.NAME, 'secondLangDegree')
    if ele.is_displayed():
        wait(1)
        if dropdown_test(ele) == 0:
            for i in range(0, 2):
                pass_counter += 1
                writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
                writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 2
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 2
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box FirstName-------------------------------
current_row = 42
try:
    ele = driver.find_element(By.NAME, 'secondLangAwardedBy')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        if len(ele.get_attribute('value')) == 0:
            ele.send_keys("unisecondLangAwardedBy")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------File upload qualification Selection-------------------------------
current_row = 44
try:
    ele = driver.find_element(By.ID, 'myfile2')
    if ele.is_displayed():
        wait(1)
        ele.send_keys(pdf_path)
        for i in range(0, 2):
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------checkboxSelection-------------------------------
current_row = 46
try:
    ele = driver.find_element(By.ID, "dohave")
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

body = driver.find_element(By.CSS_SELECTOR, 'body')
body.send_keys(Keys.PAGE_DOWN)
wait(2)

# ------------------------- dropdown - Selection -------------------------------
current_row = 47
try:
    ele = driver.find_element(By.NAME, "certificationL1L2")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box FirstName-------------------------------
current_row = 48
try:
    ele = driver.find_element(By.XPATH, '//tbody/tr[1]/td[2]/div[1]/input[1]')
    if input_box(ele, 'unisecondLangAwardedBy') == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 49
try:
    ele = driver.find_element(By.NAME, "durationCertificationL1L2")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 50
try:
    ele = driver.find_element(By.NAME, "modeCertificationL1L2")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------File upload certificate Selection-------------------------------
current_row = 51
try:
    ele = driver.find_element(By.ID, 'myfile3')
    if ele.is_displayed():
        wait(1)
        ele.send_keys(pdf_path)
        for i in range(0, 2):
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row + i, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------checkboxSelection-------------------------------
current_row = 53
try:
    ele = driver.find_element(By.ID, "experience")
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

body = driver.find_element(By.CSS_SELECTOR, 'body')
body.send_keys(Keys.PAGE_DOWN)
wait(2)

# ------------------------- dropdown - Selection -------------------------------
current_row = 54
try:
    ele = driver.find_element(By.NAME, "expFirstLang")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 55
try:
    ele = driver.find_element(By.NAME, "expSecondLang")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 56
try:
    ele = driver.find_element(By.NAME, "scriptWrittenFirstLang")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 57
try:
    ele = driver.find_element(By.NAME, 'scriptWrittenFirstLangLink')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 58
try:
    ele = driver.find_element(By.NAME, "scriptWrittenSecondLang")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 59
try:
    ele = driver.find_element(By.NAME, 'scriptWrittenSecondLangLink')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 60
try:
    ele = driver.find_element(By.NAME, "noScriptTransFromL1toL2")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 61
try:
    ele = driver.find_element(By.NAME, 'noScriptTransFromL1toL2Link')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 62
try:
    ele = driver.find_element(By.NAME, "noScriptTransFromL2toL1")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 63
try:
    ele = driver.find_element(By.NAME, 'noScriptTransFromL2toL1Link')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 64
try:
    ele = driver.find_element(By.NAME, "noArticleTransFromL1toL2")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 65
try:
    ele = driver.find_element(By.NAME, 'noArticleTransFromL1toL2Link')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 66
try:
    ele = driver.find_element(By.NAME, "noArticleTransFromL2toL1")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------input box -------------------------------
current_row = 67
try:
    ele = driver.find_element(By.NAME, 'noArticleTransFromL2toL1Link')
    if ele.is_displayed():
        ele.send_keys("#@$%!")
        wait(1)
        ele.send_keys("unisecondLangAwardedBy")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------checkboxSelection-------------------------------
current_row = 68
try:
    ele = driver.find_element(By.XPATH,
                              "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/form[1]/div[3]/input[1]")
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

body = driver.find_element(By.CSS_SELECTOR, 'body')
body.send_keys(Keys.PAGE_UP)
wait(2)
body.send_keys(Keys.PAGE_UP)
wait(2)
body.send_keys(Keys.PAGE_UP)
wait(5)
# -------------------------left info option Selection-------------------------------
current_row = 70
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container div.accordion:nth-child(1) div.accordion-item:nth-child(1) h2.accordion-header > button.accordion-button.collapsed")
    if ele.is_displayed():
        ele2 = driver.find_element(By.XPATH, "// div[contains(text(), 'Mynextfilm.com endeavours to deliver directly usab')]")
        if ele2.is_displayed():
            ele.click()
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------left info option Selection-------------------------------
current_row = 72
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container div.accordion:nth-child(1) div.accordion-item:nth-child(2) h2.accordion-header > button.accordion-button")
    if ele.is_displayed():
        ele.click()
        ele2 = driver.find_element(By.XPATH, "//li[contains(text(),'The edited version is sent to the user. A user acc')]")
        if ele2.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------left info option Selection-------------------------------
current_row = 74
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container div.accordion:nth-child(1) div.accordion-item:nth-child(3) h2.accordion-header > button.accordion-button.collapsed")
    if ele.is_displayed():
        ele.click()
        ele2 = driver.find_element(By.XPATH, "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]")
        if ele2.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------left info option Selection-------------------------------
current_row = 76
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container div.accordion:nth-child(1) div.accordion-item:nth-child(4) h2.accordion-header > button.accordion-button.collapsed")
    if ele.is_displayed():
        ele.click()
        ele2 = driver.find_element(By.XPATH, "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]")
        if ele2.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------mail option Selection-------------------------------
current_row = 77
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'info@mynextfilm.com')]")
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele = driver.find_element(By.ID, 'email_subject')
        if ele.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
writeData(inputFileName, inputSheetName, r + current_row, c + current_column - 1, 'yes')
wait(1)

# -------------------------mail option Selection-------------------------------
print(f"pass = {pass_counter}")
# writeData(inputFileName, inputSheetName, r + 1, c + 18, pass_counter)
print(f"fail = {fail_counter}")
# writeData(inputFileName, inputSheetName, r + 2, c + 18, fail_counter)
print(f"Total = {pass_counter + fail_counter}")
# writeData(inputFileName, inputSheetName, r + 3, c + 18, pass_counter + fail_counter)
