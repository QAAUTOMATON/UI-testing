import os
import time
from datetime import date

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'JoinAsEng'
inputFileName = 'OpportunitiesWithMNF.xlsx'
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"
day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()

# =======================
def wait(n):
    return time.sleep(n / 2)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)

# ===================================================
current_row = 5
current_column = 18
# -------------------------Logging in-------------------------------
login()

current_row = 5
try:
    ele = driver.find_element(By.CSS_SELECTOR, 'body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(2) div.dropdown > button.dropbtn1.opport-w-m')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 6
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'Join as Engineer')]")
    if ele.is_displayed():
        ele.click()
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 7
try:
    ele = driver.find_element(By.XPATH, "//strong[contains(text(),'Data Science Lead')]")
    if ele.is_displayed():
        ele.click()
        wait(2)
        ele2 = driver.find_element(By.XPATH, "//body/main[1]/section[1]/div[1]/div[1]/div[1]/div[1]/div[1]/p[2]/a[1]")
        if ele2.is_displayed():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 8
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if not(ele2.is_displayed()):
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 9
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            wait(2)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 10
try:
    ele = driver.find_element(By.NAME, 'firstname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'firstName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 11
try:
    ele = driver.find_element(By.NAME, 'lastname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'lastName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 12
try:
    ele = driver.find_element(By.NAME, 'email')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("autotest123@qwe.in")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 15
try:
    ele = driver.find_element(By.NAME, 'contact')
    if ele.is_displayed():
        ele.click()
        if input_box_number(ele, '4204204201') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 16
try:
    ele = driver.find_element(By.NAME, 'location')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("hNo 234 City Bangalore - 353535")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 17
try:
    ele = driver.find_element(By.NAME, 'resume')
    ele.click()
    if ele.is_displayed():
        ele.click()
        wait(5)
        os.startfile('sendFilePDF.exe')
        wait(10)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 18
try:
    ele = driver.find_element(By.NAME, 'profileLinkedin')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.linkedin.com/signup")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 19
try:
    ele = driver.find_element(By.NAME, 'portfolioLink')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 20
try:
    ele = driver.find_element(By.NAME, 'ExtraInfo')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("ExtraInfo is  site link https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

pageup()

current_row = 22
try:
    ele = driver.find_element(By.XPATH, "//strong[contains(text(),'Senior Python/Django Developer')]")
    if ele.is_displayed():
        ele.click()
        wait(2)
        ele2 = driver.find_element(By.XPATH, "//body/main[1]/section[1]/div[1]/div[1]/div[2]/div[1]/div[1]/p[2]/a[1]")
        if ele2.is_displayed():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 23
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if not(ele2.is_displayed()):
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 24
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            wait(2)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 25
try:
    ele = driver.find_element(By.NAME, 'firstname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'firstName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 26
try:
    ele = driver.find_element(By.NAME, 'lastname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'lastName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 27
try:
    ele = driver.find_element(By.NAME, 'email')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("autotest123@qwe.in")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox no----------------------------------------------------
current_row = 30
try:
    ele = driver.find_element(By.NAME, 'contact')
    if ele.is_displayed():
        ele.click()
        if input_box_number(ele, '4204204201') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 31
try:
    ele = driver.find_element(By.NAME, 'location')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("hNo 234 City Bangalore - 353535")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 32
try:
    ele = driver.find_element(By.NAME, 'resume')
    if ele.is_displayed():
        ele.click()
        wait(5)
        os.startfile('sendFilePDF.exe')
        wait(10)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 33
try:
    ele = driver.find_element(By.NAME, 'profileLinkedin')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.linkedin.com/signup")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 34
try:
    ele = driver.find_element(By.NAME, 'portfolioLink')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 35
try:
    ele = driver.find_element(By.NAME, 'ExtraInfo')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("ExtraInfo is  site link https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 37
try:
    ele = driver.find_element(By.XPATH, "//strong[contains(text(),'Senior Frontend Developer')]")
    if ele.is_displayed():
        ele.click()
        wait(2)
        ele2 = driver.find_element(By.XPATH, "//body/main[1]/section[1]/div[1]/div[1]/div[3]/div[1]/div[1]/p[2]/a[1]")
        if ele2.is_displayed():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

pageup()

current_row = 38
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if not(ele2.is_displayed()):
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 39
try:
    if ele.is_displayed():
        ele.click()
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            wait(2)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 40
try:
    ele = driver.find_element(By.NAME, 'firstname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'firstName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 41
try:
    ele = driver.find_element(By.NAME, 'lastname')
    if ele.is_displayed():
        ele.click()
        if input_box(ele, 'lastName') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 42
try:
    ele = driver.find_element(By.NAME, 'email')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("autotest123@qwe.in")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox name----------------------------------------------------
current_row = 45
try:
    ele = driver.find_element(By.NAME, 'contact')
    if ele.is_displayed():
        ele.click()
        if input_box_number(ele, '4204204201') == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox mail----------------------------------------------------
current_row = 46
try:
    ele = driver.find_element(By.NAME, 'location')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("hNo 234 City Bangalore - 353535")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------file upload----------------------------------------------------
current_row = 47
try:
    ele = driver.find_element(By.NAME, 'resume')
    if ele.is_displayed():
        ele.click()
        os.startfile('sendFilePDF.exe')
        wait(10)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox link----------------------------------------------------
current_row = 48
try:
    ele = driver.find_element(By.NAME, 'profileLinkedin')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.linkedin.com/signup")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox link----------------------------------------------------
current_row = 49
try:
    ele = driver.find_element(By.NAME, 'portfolioLink')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# --------------textbox link----------------------------------------------------
current_row = 50
try:
    ele = driver.find_element(By.NAME, 'ExtraInfo')
    if ele.is_displayed():
        ele.click()
        ele.send_keys("ExtraInfo is  site link https://www.portfolioLink.com")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

