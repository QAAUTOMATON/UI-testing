import time
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

inputSheetName = 'convertedPPT'
inputFileName = 'conversionStudio.xlsx'
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()


# =======================
def wait(n):
    return time.sleep(n / 6)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element):
    if element.is_displayed():
        element.clear()
        element.send_keys("@#$%!@#$%^&*")
        element.send_keys("QWERTYUqwerty")
        element.send_keys("1234567890")
        wait(1)
        return 1



def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)

def input_limit(element):
    count = 0
    for i in range(0, 62):
        element.send_keys("abcdefghijklmnop")
        count += 16
    if count < 1000:
        return 0
    else:
        return 1

# ===================================================
current_row = 6
current_column = 18
# -------------------------Logging in-------------------------------
login()

# navigation to top navbar
current_row = 6
try:
    ele = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[6]/div/button')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        # navigation to top navbar options
        ele2 = driver.find_element(By.XPATH, "//a[contains(text(),'View Your Conversions')]")
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# opening ppt option
current_row = 7
try:
    ele = driver.find_element(By.ID, 't_ppt')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# opening 3 dot option
current_row = 8
try:
    ele = driver.find_element(By.XPATH, '//*[@id="div_ppt"]/div[1]/div/div/div[1]/div/div[1]/div/a/i')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(2)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 9
try:
    ele = driver.find_element(By.XPATH, '//*[@id="div_ppt"]/div[1]/div/div/div[1]/div/div[1]/div/a/i')
    if ele.is_displayed():
        ele.click()
        print(current_row)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(10)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 18
try:
    ele = driver.find_element(By.XPATH, '/html/body/div[8]/div/div[3]/div/div/div[1]/div[3]/div/div/div[1]/div/div[1]/div/button')
    if ele.is_displayed():
        ele.click()
        ele2 = driver.find_element(By.ID, 'feed_btn2')
        if ele.is_displayed():
            ele.click()
            wait(1)
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

try:
    ele = driver.find_element(By.XPATH, "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/button[2]")
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        print(current_row)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 16

try:
    ele = driver.find_element(By.ID, "fed_text2")
    if ele.is_displayed():
        ele.click()
        if input_box(ele) == 1:
            wait(2)
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 17

try:
    ele = driver.find_element(By.ID, "error")
    if ele.is_displayed():
        ele.click()
        if input_box(ele) == 1:
            wait(2)
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 18

try:
    ele = driver.find_element(By.ID, "improvement")
    if ele.is_displayed():
        ele.click()
        if input_box(ele) == 1:
            wait(2)
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 13

try:
    ele1 = driver.find_element(By.ID, "st12")
    ele2 = driver.find_element(By.ID, "st22")
    ele3 = driver.find_element(By.ID, "st32")
    ele4 = driver.find_element(By.ID, "st42")
    ele5 = driver.find_element(By.ID, "st52")
    star_count = 0
    if ele1.is_displayed():
        ele1.click()
        wait(2)
        star_count += 1
        ele2.click()
        wait(2)
        star_count += 1
        ele3.click()
        wait(2)
        star_count += 1
        ele4.click()
        wait(2)
        star_count += 1
        ele5.click()
        wait(2)
        star_count += 1
    if star_count == 5:
        print(current_row)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 14
print(current_row)

try:
    ele1 = driver.find_element(By.ID, "accurate1")
    ele2 = driver.find_element(By.ID, "accurate2")
    ele3 = driver.find_element(By.ID, "accurate3")
    ele4 = driver.find_element(By.ID, "accurate4")
    ele5 = driver.find_element(By.ID, "accurate5")
    star_count = 0
    if ele1.is_displayed():
        ele1.click()
        wait(2)
        star_count += 1
        ele2.click()
        wait(2)
        star_count += 1
        ele3.click()
        wait(2)
        star_count += 1
        ele4.click()
        wait(2)
        star_count += 1
        ele5.click()
        wait(2)
        star_count += 1
    if star_count == 5:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 15
print(current_row)

try:
    ele1 = driver.find_element(By.ID, "language1")
    ele2 = driver.find_element(By.ID, "language2")
    ele3 = driver.find_element(By.ID, "language3")
    ele4 = driver.find_element(By.ID, "language4")
    ele5 = driver.find_element(By.ID, "language5")
    star_count = 0
    if ele1.is_displayed():
        ele1.click()
        wait(2)
        star_count += 1
        ele2.click()
        wait(2)
        star_count += 1
        ele3.click()
        wait(2)
        star_count += 1
        ele4.click()
        wait(2)
        star_count += 1
        ele5.click()
        wait(2)
        star_count += 1
    if star_count == 5:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 19
try:
    ele = driver.find_element(By.ID, "feed_btn")
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

