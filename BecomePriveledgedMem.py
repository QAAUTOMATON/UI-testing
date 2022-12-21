import os
import time
from datetime import date

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'ClaimPrev'
inputFileName = 'BecomePriveledgedMem.xlsx'
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()


# =======================
def wait(n):
    return time.sleep(n / 4)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)


# ===================================================
current_row = 6
current_column = 18
# -------------------------Logging in-------------------------------
login()

ele = driver.find_element(By.CSS_SELECTOR,
                          "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(1) div.dropdown > button.dropbtn1.opport.claim-y-p")
ele.click()
wait(5)
ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center ul.nav.navbar-nav li.nav-item:nth-child(1) div.dropdown div.dropdown-content > a:nth-child(1)")
ele.click()

current_row = 6
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(9) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div:nth-child(6) div.main-container:nth-child(6) div.item2 form:nth-child(3) > button:nth-child(4)")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 10
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div:nth-child(7) div.main-container:nth-child(6) div.item3 form:nth-child(2) > button:nth-child(4)")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 12
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.m-div:nth-child(13) div.container-fluid div.text-center:nth-child(5) div.checkout-container div.checkout-content div:nth-child(2) div.row div.col-sm-6:nth-child(5) > button.s-btn.font-weight-bold:nth-child(2)")
    if ele.is_displayed():
        wait(5)
        ele.click()
        ele2 = driver.find_element(By.ID, "codee")
        wait(1)
        if ele2.is_selected():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 13
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'Back To Payment')]")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 14
try:
    ele = driver.find_element(By.XPATH, '//*[@id="left-component"]/div/div[2]/div[4]/form/button')
    if ele.is_displayed():
        wait(5)
        ele.click()
        wait(5)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 16
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.m-div:nth-child(13) div.container-fluid div.text-center:nth-child(5) div.checkout-container div.checkout-content div:nth-child(2) div.row div.col-sm-6:nth-child(5) > button.s-btn.font-weight-bold:nth-child(2)")
    if ele.is_displayed():
        wait(5)
        ele.click()
        ele2 = driver.find_element(By.ID, "codee")
        wait(1)
        if ele2.is_selected():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 17
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'Back To Payment')]")
    if ele.is_displayed():
        ele.click()
        wait(5)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


current_row = 28
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div:nth-child(7) div.main-container:nth-child(6) div.item3 form:nth-child(2) > button:nth-child(4)")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 32
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.m-div:nth-child(13) div.container-fluid div.text-center:nth-child(5) div.checkout-container div.checkout-content div:nth-child(2) div.row div.col-sm-6:nth-child(5) > button.s-btn.font-weight-bold:nth-child(2)")
    if ele.is_displayed():
        wait(2)
        ele.click()
        ele2 = driver.find_element(By.ID, "codee")
        wait(1)
        ele2.send_keys('invalidPromo')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(1)
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 33
try:
    ele = driver.find_element(By.ID, 'submit-button')
    if ele.is_displayed():
        ele.click()
        wait(5)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 29
try:
    ele = driver.find_element(By.ID, 'contact')
    if ele.is_displayed():
        ele.click()
        ele.send_keys('9624459265')
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 30
try:
    ele = driver.find_element(By.ID, 'email')
    if ele.is_displayed():
        ele.click()
        ele.send_keys('	automationmnf123@gmail.com')
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 34
try:
    ele = driver.find_element(By.ID, 'modal-close')
    if ele.is_displayed():
        ele.click()
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)



current_row = 35
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div:nth-child(7) div.main-container:nth-child(6) div.item4 form:nth-child(4) > button:nth-child(4)")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 39
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.m-div:nth-child(13) div.container-fluid div.text-center:nth-child(5) div.checkout-container div.checkout-content div:nth-child(2) div.row div.col-sm-6:nth-child(5) > button.s-btn.font-weight-bold:nth-child(2)")
    if ele.is_displayed():
        wait(2)
        ele.click()
        ele2 = driver.find_element(By.ID, "codee")
        wait(1)
        ele2.send_keys('invalidPromo')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(1)
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 40
try:
    ele = driver.find_element(By.ID, 'submit-button')
    if ele.is_displayed():
        ele.click()
        wait(5)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 29
try:
    ele = driver.find_element(By.ID, 'contact')
    if ele.is_displayed():
        ele.click()
        ele.send_keys('9624459265')
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 30
try:
    ele = driver.find_element(By.ID, 'email')
    if ele.is_displayed():
        ele.click()
        ele.send_keys('	automationmnf123@gmail.com')
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 34
try:
    ele = driver.find_element(By.ID, 'modal-close')
    if ele.is_displayed():
        ele.click()
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


current_row = 47
try:
    ele = driver.find_element(By.CSS_SELECTOR, "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.m-div:nth-child(13) div.container-fluid div.text-center:nth-child(5) div.checkout-container div.checkout-content div:nth-child(2) div.row div.col-sm-6:nth-child(5) > button.s-btn.font-weight-bold:nth-child(2)")
    if ele.is_displayed():
        wait(2)
        ele.click()
        ele2 = driver.find_element(By.ID, "codee")
        wait(1)
        ele2.send_keys('invalidPromo')
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        wait(1)
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 48
try:
    ele = driver.find_element(By.ID, 'submit-button')
    if ele.is_displayed():
        ele.click()
        wait(5)
        if driver.current_url == "https://mynextfilm.com/pay/payment/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)