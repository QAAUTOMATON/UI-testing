import time
from datetime import date

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'GiftSubs'
inputFileName = 'BecomePriveledgedMem.xlsx'
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()


# =======================
def wait(n):
    return time.sleep(n/6)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_mail(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)


# ===================================================
current_row = 6
current_column = 18
# -------------------------Logging in-------------------------------
login()

ele = driver.find_element(By.CSS_SELECTOR,
                          "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(1) div.dropdown > button.dropbtn1.opport.claim-y-p")
ele.click()
wait(5)
ele = driver.find_element(By.ID, "giftDivMain")
ele.click()

current_row = 13
try:
    ele = driver.find_element(By.ID, "add1")
    if ele.is_displayed():
        ele.click()
        wait(2)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

wait(20)
current_row = 6
try:
    eleName = driver.find_element(By.ID, "Gt1")
    if eleName.is_displayed():
        eleName.click()
        eleName.clear()
        if input_box(eleName, 'firstname') == 0:
            print(0)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 6
try:
    eleName = driver.find_element(By.ID, "Gt1")
    if eleName.is_displayed():
        eleName.click()
        if input_box(eleName, 'firstname') == 0:
            print(0)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 7
try:
    eleMail = driver.find_element(By.ID, "G1")
    if eleMail.is_displayed():
        eleMail.click()
        if input_box_mail(eleMail, 'somemail@01.er') != 0:
            eleMail.clear()
            wait(2)
            eleMail.send_keys("somemail@01.er")
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 8
try:
    eleNum = driver.find_element(By.ID, "Gp1")
    if eleNum.is_displayed():
        eleNum.click()
        if input_box_number(eleNum, "9876675645"):
            eleNum.clear()
            wait(2)
            eleNum.send_keys("9876675645")
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 9
try:
    eleCc = driver.find_element(By.ID, "code1")
    if eleCc.is_displayed():
        eleCc.click()
        if dropdown_test(eleCc) == 0:
            Select(eleCc).select_by_visible_text("India (+91)")
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

elebody = driver.find_element(By.XPATH, '//*[@id="kt_body"]/div[6]')
current_row = 14
try:
    ele = driver.find_element(By.ID, "add1")
    if ele.is_displayed():
        eleMail.clear()
        wait(2)
        eleMail.click()
        wait(2)
        eleMail.send_keys("samplemail@mnf.ac.in")
        wait(1)
        print(1)
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 15
try:
    ele = driver.find_element(By.ID, "add1")
    if ele.is_displayed():
        eleName.send_keys("@#!")
        print(2)
        ele.click()
        wait(1)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 16
try:
    ele = driver.find_element(By.ID, "add1")
    if ele.is_displayed():
        eleName.clear()
        eleName.send_keys("myName")
        wait(1)
        ele.click()
        print(1)
        elebody.click()
        wait(1)
        if driver.find_element(By.ID, "Gt2").is_displayed():
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


# eleName.click()
driver.find_element(By.XPATH, '//*[@id="referform1"]/div[2]/div[2]/a/i').click()
wait(5)

current_row = 10
try:
    ele = driver.find_element(By.ID, "submitButton1")
    if ele.is_displayed():
        ele.click()
        wait(10)
        if driver.current_url != "https://mynextfilm.com/pay/gift/":
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 11
try:
    ele = driver.find_element(By.ID, "codee")
    if ele.is_displayed():
        ele.send_keys("code10")
        wait(1)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 12
try:
    ele = driver.find_element(By.XPATH, "//a[contains(text(),'Back To Payment')]")
    if ele.is_displayed():
        ele.click()
        wait(10)
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)