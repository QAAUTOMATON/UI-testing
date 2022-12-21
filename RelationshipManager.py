import time
from datetime import date

import os

import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'RelationshipManager'
inputFileName = 'OpportunitiesWithMNF.xlsx'

pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"

path = "C:\Program Files (x86)\chromedriver.exe"

day = date.today()

# ==========================================================================================================
driver = webdriver.Chrome(path)

url = 'https://mynextfilm.com/members-home'
driver.get(url)

driver.maximize_window()

# =======================
def wait(n):
    return time.sleep(n / 6)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def dragAndDrop(sourceEle, x, y):
    source = driver.find_element(By.ID, sourceEle)
    action = ActionChains(driver)
    action.drag_and_drop_by_offset(source, x, y).perform()
    wait(5)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


# ===================================================
current_row = 5
current_column = 18
# -------------------------Logging in-------------------------------
login()
# -------------------------Navigation to Opp With MNF option-------------------------------
current_row = 5
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              'body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(2) div.dropdown > button.dropbtn1.opport-w-m')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------Navigation to Opp With MNF: RM option-------------------------------
current_row = 6
try:
    ele = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[11]/ul/li[2]/div/div/a[2]')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------left info option Selection-------------------------------
current_row = 7
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container.my-4 div.accordion:nth-child(1) div.accordion-item:nth-child(1) h2.accordion-header > button.accordion-button")
    if ele.is_displayed():
        # element already open
        ele2 = driver.find_element(By.XPATH,
                                   "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]")
        if ele2.is_displayed():
            ele.click()
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------left info option Selection-------------------------------
current_row = 9
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container.my-4 div.accordion:nth-child(1) div.accordion-item:nth-child(2) h2.accordion-header > button.accordion-button")
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele2 = driver.find_element(By.XPATH,
                                   "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]")
        if ele2.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------left info option Selection-------------------------------
current_row = 11
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container.my-4 div.accordion:nth-child(1) div.accordion-item:nth-child(3) h2.accordion-header > button.accordion-button.collapsed")
    if ele.is_displayed():
        ele.click()
        wait(1)
        ele2 = driver.find_element(By.XPATH,
                                   "//body/div[8]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]")
        if ele2.is_displayed():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------left info option Selection-------------------------------
current_row = 13
try:
    ele = driver.find_element(By.CSS_SELECTOR,
                              "body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid div.pretty-split-pane-frame div.split-pane.vertical-percent div.split-pane-component:nth-child(1) div.row:nth-child(11) div.col-sm-4.my-5.mx-auto div.container.my-4 p:nth-child(2) span:nth-child(1) b:nth-child(1) > a:nth-child(1)")
    if ele.is_displayed():
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# skip 13
# -------------------------------submit without filling details----------------------------------
current_row = 14
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling firstname details----------------------------------
current_row = 15
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    ele2 = driver.find_element(By.NAME, 'firstName')
    if ele.is_displayed():
        ele2.clear()
        wait(1)
        ele.click()
        wait(1)
        if ele2.is_displayed():
            ele2.send_keys('firstName')
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling lastName details----------------------------------
current_row = 16
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    ele2 = driver.find_element(By.NAME, 'lastName')
    if ele.is_displayed():
        ele2.clear()
        wait(1)
        ele.click()
        wait(1)
        if ele2.is_displayed():
            ele2.send_keys('lastName')
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling mail details----------------------------------
current_row = 17
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    ele2 = driver.find_element(By.NAME, 'emailId')
    if ele.is_displayed():
        ele2.clear()
        wait(1)
        ele.click()
        wait(1)
        if ele2.is_displayed():
            ele2.send_keys("emailId@mail")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling phone details----------------------------------
current_row = 18
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    ele2 = driver.find_element(By.NAME, 'number')
    if ele.is_displayed():
        ele2.clear()
        wait(1)
        ele.click()
        wait(1)
        if ele2.is_displayed():
            ele2.send_keys("8776767880")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 19
try:
    ele = driver.find_element(By.ID, "selectNumber")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 20
try:
    ele = driver.find_element(By.NAME, "qualification")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------submit without filling details----------------------------------
current_row = 21
try:
    ele = driver.find_element(By.ID, 'submitIdd')
    ele2 = driver.find_element(By.NAME, 'qualificationAwardedBy')
    if ele.is_displayed():
        ele2.clear()
        wait(1)
        ele.click()
        wait(1)
        if ele2.is_displayed():
            ele2.send_keys("someUni")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 23
try:
    ele = driver.find_element(By.ID, "country")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 24
try:
    ele = driver.find_element(By.ID, "country")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 25
try:
    ele = driver.find_element(By.ID, "region")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column , day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

wait(1)

# ------------------------- dropdown - Selection -------------------------------
current_row = 27
try:
    ele = driver.find_element(By.ID, "city")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 29
try:
    ele = driver.find_element(By.ID, "selectNumber")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 30
try:
    ele = driver.find_element(By.ID, "secondSpokenLang")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 31
try:
    ele = driver.find_element(By.NAME, "qualification")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)


# -------------------------------textbox firstname details----------------------------------
current_row = 32
try:
    ele = driver.find_element(By.NAME, 'firstName')
    if input_box(ele, 'firstname') == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling lastName details----------------------------------
current_row = 33
try:
    ele = driver.find_element(By.NAME, 'lastName')
    if input_box(ele, 'lastName') == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------submit without filling mail details----------------------------------
current_row = 34
try:
    ele = driver.find_element(By.NAME, 'emailId')
    if ele.is_displayed():
        ele.clear()
        wait(1)
        ele.send_keys("emailId@mail.com")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box Number details----------------------------------
current_row = 35
try:
    ele = driver.find_element(By.NAME, 'number')
    if ele.is_displayed():
        ele.clear()
        wait(1)
        ele.send_keys("4565987605")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box other language details----------------------------------
current_row = 36
try:
    ele = driver.find_element(By.ID, 'selectNumber')
    if ele.is_displayed():
        Select(ele).select_by_visible_text("Others")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        wait(1)
        ele2 = driver.find_element(By.ID, 'hide14')
        if ele2.is_displayed():
            wait(1)
            ele.send_keys("someLanguage")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box other language details----------------------------------
current_row = 37
try:
    ele = driver.find_element(By.NAME, 'qualification')
    if ele.is_displayed():
        Select(ele).select_by_visible_text("Others")
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        wait(1)
        ele2 = driver.find_element(By.ID, 'hide13')
        if ele2.is_displayed():
            wait(1)
            ele2.send_keys("someLanguage")
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row+1, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 39
try:
    ele = driver.find_element(By.NAME, 'qualificationAwardedBy')
    if input_box(ele, "qualificationAwardedBy") == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 40
try:
    ele = driver.find_element(By.NAME, 'videolink')
    if ele.is_displayed():
        ele.send_keys("https://mynextfilm.com/pay/payment/")
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 41
try:
    ele = driver.find_element(By.NAME, 'motherTounge')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 42
try:
    ele = driver.find_element(By.NAME, 'nativLanguage')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 43
try:
    ele = driver.find_element(By.ID, 'qualifiedInParticularLanguage')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 44
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box----------------------------------
current_row = 45
try:
    ele = driver.find_element(By.NAME, 'mediumOfInstructionL1')
    if dropdown_test(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 46
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction')
    if ele.is_selected():
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 47
try:
    ele = driver.find_element(By.ID, 'mediumOfInstruction')
    ele2 = driver.find_element(By.NAME, 'nativLanguage')
    ele3 = driver.find_element(By.ID, 'qualifiedInParticularLanguage')
    if ele.is_selected() and ele2.is_selected() and ele3.is_selected():
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)


# -------------------------------checkbox box ----------------------------------
current_row = 50
try:
    ele = driver.find_element(By.ID, 'ScriptWriting')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 51
try:
    ele = driver.find_element(By.ID, 'Acting')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 52
try:
    ele = driver.find_element(By.ID, 'Directing')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 53
try:
    ele = driver.find_element(By.ID, 'ProducingFilms')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 54
try:
    ele = driver.find_element(By.ID, 'Editing')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 55
try:
    ele = driver.find_element(By.ID, 'Animation')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)




# -------------------------------checkbox box other language details----------------------------------
current_row = 59
try:
    ele = driver.find_element(By.ID, 'Aspirant')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 60
try:
    ele = driver.find_element(By.ID, 'Enthusist')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 61
try:
    ele = driver.find_element(By.ID, 'Amateur')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 62
try:
    ele = driver.find_element(By.ID, 'Professional')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 63
try:
    ele = driver.find_element(By.ID, 'Expert')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box other language details----------------------------------
current_row = 64
try:
    ele = driver.find_element(By.ID, 'myfile')
    if ele.is_displayed():
        wait(1)
        ele.click()
        wait(2)
        os.startfile("sendFilePDF.exe")
        wait(4)
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 67
try:
    ele = driver.find_element(By.XPATH, '//*[@id="left-component"]/div/div/div[2]/div/div[1]/div[6]/input')
    if checkbox(ele) == 0:
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 68
try:
    ele.click()
    driver.find_element(By.XPATH, "//thead/tr/th[1]").click()
    if not(ele.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
print("done")
ele.click()
pagedown()
# -------------------------------checkbox box ----------------------------------
ele12 = driver.find_element(By.XPATH, "//input[@id='morning']")
ele13 = driver.find_element(By.XPATH, '//tbody/tr[1]/td[3]/span[1]/input[1]')
ele14 = driver.find_element(By.XPATH, '//tbody/tr[1]/td[4]/span[1]/input[1]')

ele21 = driver.find_element(By.XPATH, '//tbody/tr[2]/td[1]/span[1]/input[1]')
ele22 = driver.find_element(By.XPATH, '//tbody/tr[2]/td[2]/span[1]/input[1]')
ele23 = driver.find_element(By.XPATH, '//tbody/tr[2]/td[3]/span[1]/input[1]')
ele24 = driver.find_element(By.XPATH, '//tbody/tr[2]/td[4]/span[1]/input[1]')
ele31 = driver.find_element(By.XPATH, '//tbody/tr[3]/td[1]/span[1]/input[1]')
ele32 = driver.find_element(By.XPATH, '//tbody/tr[3]/td[2]/span[1]/input[1]')
ele33 = driver.find_element(By.XPATH, '//tbody/tr[3]/td[3]/span[1]/input[1]')
ele34 = driver.find_element(By.XPATH, '//tbody/tr[3]/td[4]/span[1]/input[1]')
ele41 = driver.find_element(By.XPATH, '//tbody/tr[4]/td[1]/span[1]/input[1]')
ele42 = driver.find_element(By.XPATH, '//tbody/tr[4]/td[2]/span[1]/input[1]')
ele43 = driver.find_element(By.XPATH, '//tbody/tr[4]/td[3]/span[1]/input[1]')
ele44 = driver.find_element(By.XPATH, '//tbody/tr[4]/td[4]/span[1]/input[1]')
ele51 = driver.find_element(By.XPATH, '//tbody/tr[5]/td[1]/span[1]/input[1]')
ele52 = driver.find_element(By.XPATH, '//tbody/tr[5]/td[2]/span[1]/input[1]')
ele53 = driver.find_element(By.XPATH, '//tbody/tr[5]/td[3]/span[1]/input[1]')
ele54 = driver.find_element(By.XPATH, '//tbody/tr[5]/td[4]/span[1]/input[1]')
ele61 = driver.find_element(By.XPATH, '//tbody/tr[6]/td[1]/span[1]/input[1]')
ele62 = driver.find_element(By.XPATH, '//tbody/tr[6]/td[2]/span[1]/input[1]')
ele63 = driver.find_element(By.XPATH, '//tbody/tr[6]/td[3]/span[1]/input[1]')
ele64 = driver.find_element(By.XPATH, '//tbody/tr[6]/td[4]/span[1]/input[1]')
ele71 = driver.find_element(By.XPATH, '//tbody/tr[7]/td[1]/span[1]/input[1]')
ele72 = driver.find_element(By.XPATH, '//tbody/tr[7]/td[2]/span[1]/input[1]')
ele73 = driver.find_element(By.XPATH, '//tbody/tr[7]/td[3]/span[1]/input[1]')
ele74 = driver.find_element(By.XPATH, '//tbody/tr[7]/td[4]/span[1]/input[1]')
ele81 = driver.find_element(By.XPATH, '//tbody/tr[8]/td[1]/span[1]/input[1]')
ele82 = driver.find_element(By.XPATH, '//tbody/tr[8]/td[2]/span[1]/input[1]')
ele83 = driver.find_element(By.XPATH, '//tbody/tr[8]/td[3]/span[1]/input[1]')
ele84 = driver.find_element(By.XPATH, '//tbody/tr[8]/td[4]/span[1]/input[1]')

# -------------------------------checkbox box ----------------------------------
current_row = 69
try:
    if checkbox(ele12) == 0:
        if ele22.is_selected() and ele52.is_selected() and ele82.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 71
try:
    ele22.click()
    if not(ele12.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 72
try:
    if checkbox(ele13) == 0:
        if ele33.is_selected() and ele53.is_selected() and ele73.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 74
try:
    ele43.click()
    if not(ele13.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 75
try:
    if checkbox(ele14) == 0:
        if ele24.is_selected() and ele54.is_selected() and ele84.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)


# -------------------------------checkbox box ----------------------------------
current_row = 77
try:
    ele64.click()
    if not(ele14.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 78
try:
    if checkbox(ele21) == 0:
        if ele22.is_selected() and ele23.is_selected() and ele24.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 80
try:
    ele24.click()
    if not(ele21.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 81
try:
    if checkbox(ele31) == 0:
        if ele32.is_selected() and ele33.is_selected() and ele34.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 83
try:
    ele34.click()
    if not(ele31.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 84
try:
    if checkbox(ele41) == 0:
        if ele42.is_selected() and ele43.is_selected() and ele44.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 86
try:
    ele43.click()
    if not(ele41.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
# -------------------------------checkbox box ----------------------------------
current_row = 87
try:
    if checkbox(ele51) == 0:
        if ele52.is_selected() and ele53.is_selected() and ele54.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 89
try:
    ele52.click()
    if not(ele51.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
# -------------------------------checkbox box ----------------------------------
current_row = 90
try:
    if checkbox(ele61) == 0:
        if ele62.is_selected() and ele63.is_selected() and ele64.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 92
try:
    ele62.click()
    if not(ele61.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
# -------------------------------checkbox box ----------------------------------
current_row = 93
try:
    if checkbox(ele71) == 0:
        if ele72.is_selected() and ele73.is_selected() and ele74.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 95
try:
    ele73.click()
    if not(ele71.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
# -------------------------------checkbox box ----------------------------------
current_row = 96
try:
    if checkbox(ele81) == 0:
        if ele82.is_selected() and ele83.is_selected() and ele84.is_selected():
            pass_counter += 1
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 98
try:
    ele84.click()
    if not(ele81.is_selected()):
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
        writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 101
try:
    ele2= driver.find_element(By.ID, 'left-component')
    if ele.is_selected():
        dragAndDrop("left-component", 0, 100)
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

