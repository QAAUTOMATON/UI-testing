import time
from datetime import date

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl


# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'CollabrateWithMNF'
inputFileName = 'OpportunitiesWithMNF.xlsx'

day = date.today()
# ==========================================================================================================
path = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()

# =======================
def wait(n):
    return time.sleep(n / 2)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(2)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)

# ===================================================
current_row = 5
current_column = 18
# -------------------------Logging in-------------------------------
login()
current_row = 5
try:
    ele = driver.find_element(By.CSS_SELECTOR, 'body.quick-panel-right.demo-panel-right.offcanvas-right.header-fixed.header-mobile-fixed.aside-enabled.aside-static:nth-child(2) div.d-flex.flex-row.flex-column-fluid.page:nth-child(8) div.d-flex.flex-column.flex-row-fluid nav.navbar.navbar-expand-lg.navbar-light.bg-white.justify-content-space-between:nth-child(1) div.collapse.navbar-collapse ul.navbar-nav.w-100.justify-content-center li:nth-child(11) ul.nav.navbar-nav li.nav-item:nth-child(2) div.dropdown > button.dropbtn1.opport-w-m')
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------Navigation to Opp With MNF: Collaborate option-------------------------------
current_row = 7
try:
    ele = driver.find_element(By.XPATH, "//a[@id='collaborateDivMain']")
    if ele.is_displayed():
        ele.click()
        pass_counter += 1
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
        fail_counter += 1
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 8
try:
    ele = driver.find_element(By.ID, 'subject')
    ele.click()
    if input_box(ele, "this is a sample statement to test thhe functioning of the textbox") == 0:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------checkbox box ----------------------------------
current_row = 10
try:
    ele = driver.find_element(By.NAME, "checke")
    ele.click()
    checkbox(ele)
    if not(ele.is_selected()):
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 11
try:
    ele = driver.find_element(By.ID, 'name')
    if input_box(ele, "firstnameLastName") == 0:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 12
try:
    ele = driver.find_element(By.ID, 'Email')
    if input_box(ele, "samplemail@mail.com") == 0:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# ------------------------- dropdown - Selection -------------------------------
current_row = 13
try:
    ele = driver.find_element(By.ID, "cars")
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column + 1, day)
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
    fail_counter += 1
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

# -------------------------------input box ----------------------------------
current_row = 14
try:
    ele = driver.find_element(By.ID, 'lname')
    if input_box_number(ele, "6453526225") == 0:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

# -------------------------------button ----------------------------------
current_row = 15
try:
    ele = driver.find_element(By.XPATH, "//b[contains(text(),'the o')]")
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'Pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

