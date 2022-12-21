import os
import time
from datetime import date

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl as xl

# ==========================Initialize initial Counters===========================
# -----row and colum reference-------------
c = 0
r = 0

pass_counter = 0
fail_counter = 0
total_test_counter = 0

inputSheetName = 'videoUpload'
inputFileName = "viewer'sLounge.xlsx"
pdf_path = "C://Users/kz/OneDrive/Desktop/Automation/drive_backup/sampleupload files/ThePremonition.pdf"
vidpath = "C://Users/kz/Videos/Captures/Circuit design Epic Gaaris _ Tinkercad â€” Mozilla Firefox 2020-12-11 17-37-29.mp4"
reason = "this video beautifully explains the basics of color grading yet diving deep enough so you always learn something new."

day = date.today()
# ==========================================================================================================
path = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
url = 'https://mynextfilm.com/members-home'
driver.get(url)
driver.maximize_window()


# =======================
def wait(n):
    return time.sleep(n / 6)


def readData(file, sheetname, rownum, columnnum):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetname, rownum, columnnum, data):
    wb = xl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    wb.save(file)


def updateStatus(file, sheetname, rownum, columnnum, date):
    writeData(file, sheetname, r + rownum, c + columnnum + 1, date)
    writeData(file, sheetname, r + rownum, c + columnnum - 1, 'yes')
    wait(1)


# -----------------------test 1 log in----------------------------
def login():
    driver.find_element(By.NAME, 'username').send_keys('Automation1@gmail.com')
    wait(1)
    driver.find_element(By.ID, 'password').send_keys('Automatiomn1')
    wait(1)
    driver.find_element(By.ID, 'kt_login_signin_submit').click()
    wait(1)


# -------------------------------------------------------------------------
def dropdown_test(element):
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    wait(1)
    Select(element).select_by_index(0)
    wait(1)
    Select(element).select_by_index(1)
    return 0


# -------------------------------------------------------------------------
def word_limit_test(element, word_limit):
    word = "0123thisIsSampleTextToTestWordLimitOfGivenInputBox"
    ele = element
    ele.send_keys(word)
    if len(ele.get_attribute('value')) == word_limit:
        flag = 1
    else:
        flag = 0
    return flag


def input_box(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def input_box_number(element, keys):
    if element.is_displayed():
        element.clear()
        element.send_keys("#@$%!ABab")
        element.send_keys(keys)
        wait(1)
        if ele.get_attribute('value') == keys:
            return 0
        else:
            return 1
    else:
        return 1


def checkbox(element):
    if element.is_displayed():
        element.click()
        wait(1)
        element.click()
        wait(1)
        element.click()
        return 0
    else:
        return 1


def pagedown():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_DOWN)
    wait(2)


def pageup():
    body = driver.find_element(By.CSS_SELECTOR, 'body')
    body.send_keys(Keys.PAGE_UP)
    wait(2)

def input_limit(element):
    count = 0
    for i in range(0, 62):
        element.send_keys("abcdefghijklmnop")
        count += 16
    if count < 1000:
        return 0
    else:
        return 1

# ===================================================
current_row = 6
current_column = 18
# -------------------------Logging in-------------------------------
login()

current_row = 5
try:
    ele = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[9]/div/button')
    if ele.is_displayed():
        ele.click()
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
        ele2 = driver.find_element(By.XPATH, '//*[@id="navbarSupportedContent"]/ul/li[9]/div/div/a[1]')
        wait(2)
        if ele2.is_displayed():
            ele2.click()
            writeData(inputFileName, inputSheetName, r + current_row + 1, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
updateStatus(inputFileName, inputSheetName, current_row + 1, current_column, day)

current_row = 7
try:
    ele = driver.find_element(By.ID, 'video_upload')
    if ele.is_displayed():
        ele.click()
        if ele.is_displayed():
            ele.send_keys(vidpath)
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 8
try:
    ele = driver.find_element(By.ID, 'formFileSm')
    if ele.is_displayed():
        ele.click()
        if ele.is_displayed():
            ele.send_keys("tinkercad")
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 9
try:
    ele = driver.find_element(By.ID, 'instruction_id')
    if ele.is_displayed():
        ele.click()
        ele.send_keys(reason)
        print(current_row)
        writeData(inputFileName, inputSheetName, r + current_row , c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)

current_row = 10
try:
    ele = driver.find_element(By.NAME, 'category')
    if ele.is_displayed():
        ele.click()
        if dropdown_test(ele) == 0:
            print(current_row)
            writeData(inputFileName, inputSheetName, r + current_row , c + current_column, 'pass')
    else:
        writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'fail')
except:
    writeData(inputFileName, inputSheetName, r + current_row, c + current_column, 'N/A')
updateStatus(inputFileName, inputSheetName, current_row, current_column, day)
